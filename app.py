from fastapi import FastAPI, HTTPException, Depends
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import Response, HTMLResponse, FileResponse
import secrets
import string
from datetime import datetime, timedelta
from pydantic import BaseModel
from typing import Optional, List
import csv
from io import StringIO, BytesIO
from reportlab.lib.pagesizes import A4, landscape
from reportlab.pdfgen import canvas
from reportlab.lib.utils import ImageReader
from reportlab.lib.colors import HexColor
import os
from sqlalchemy import create_engine, Column, Integer, String, DateTime, Text, Date, text
from sqlalchemy.ext.declarative import declarative_base
from sqlalchemy.orm import sessionmaker, Session

# ============ SMS Configuration ============
# For testing, we'll simulate SMS. To enable real SMS:
# 1. Sign up at twilio.com
# 2. Get Account SID, Auth Token, and Twilio Phone Number
# 3. Uncomment the twilio import and code below

TWILIO_ACCOUNT_SID = os.environ.get('TWILIO_ACCOUNT_SID', '')
TWILIO_AUTH_TOKEN = os.environ.get('TWILIO_AUTH_TOKEN', '')
TWILIO_PHONE_NUMBER = os.environ.get('TWILIO_PHONE_NUMBER', '')

def send_sms(phone_number: str, message: str):
    """Send SMS using Twilio (simulated if credentials not set)"""
    if not TWILIO_ACCOUNT_SID or not TWILIO_AUTH_TOKEN or not TWILIO_PHONE_NUMBER:
        print(f"📱 [SIMULATED SMS] To: {phone_number}")
        print(f"📱 Message: {message}")
        return {"simulated": True, "message": "SMS simulated (no Twilio credentials)"}
    
    try:
        from twilio.rest import Client
        client = Client(TWILIO_ACCOUNT_SID, TWILIO_AUTH_TOKEN)
        message = client.messages.create(
            body=message,
            from_=TWILIO_PHONE_NUMBER,
            to=phone_number
        )
        return {"sid": message.sid, "status": "sent"}
    except Exception as e:
        print(f"❌ SMS error: {e}")
        return {"error": str(e), "status": "failed"}

# ============ FORCE POSTGRESQL CONNECTION ============
DATABASE_URL = os.environ.get('DATABASE_URL')
if not DATABASE_URL:
    raise Exception("❌ DATABASE_URL environment variable is not set! Cannot connect to PostgreSQL.")
else:
    print(f"✅ Found DATABASE_URL: {DATABASE_URL[:30]}...")

if DATABASE_URL.startswith('postgresql'):
    engine = create_engine(DATABASE_URL)
    print("✅ Using PostgreSQL engine")
else:
    raise Exception(f"❌ Invalid DATABASE_URL format: {DATABASE_URL[:20]}... Must start with 'postgresql'")

try:
    with engine.connect() as conn:
        conn.execute(text("SELECT 1"))
        print("✅ PostgreSQL connection successful!")
except Exception as e:
    print(f"❌ PostgreSQL connection failed: {e}")
    raise e

SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
Base = declarative_base()

# ============ DATABASE MODELS ============
class Registration(Base):
    __tablename__ = "registrations"
    
    id = Column(Integer, primary_key=True, index=True)
    guest_id = Column(String(20), unique=True, index=True, nullable=False)
    full_name = Column(String(100), nullable=False)
    gender = Column(String(10), nullable=False)
    phone = Column(String(20), nullable=False)
    email = Column(String(100), unique=True, nullable=False)
    date_of_birth = Column(Date, nullable=True)
    organization = Column(String(100), nullable=True)
    job_title = Column(String(100), nullable=True)
    city = Column(String(100), nullable=True)
    category = Column(String(20), nullable=False)
    registered_at = Column(DateTime, default=datetime.utcnow)
    checked_in = Column(Integer, default=0)
    checked_in_at = Column(DateTime, nullable=True)
    sms_sent = Column(Integer, default=0)
    sms_sent_at = Column(DateTime, nullable=True)

# Create tables
try:
    Base.metadata.create_all(bind=engine)
    print("✅ Database tables created/verified")
except Exception as e:
    print(f"❌ Database creation error: {e}")

# ============ FASTAPI APP ============
app = FastAPI()

# CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ============ DATABASE DEPENDENCY ============
def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()

# ============ HELPER FUNCTIONS ============
def generate_guest_id():
    prefix = "ERA75"
    suffix = ''.join(secrets.choice(string.ascii_uppercase + string.digits) for _ in range(6))
    return f"{prefix}-{suffix}"

# ============ MODELS ============
class RegistrationCreate(BaseModel):
    full_name: str
    gender: str
    phone: str
    email: str
    date_of_birth: Optional[str] = None
    organization: Optional[str] = None
    job_title: Optional[str] = None
    city: Optional[str] = None
    category: str

# ============ ROUTES ============
@app.get("/", response_class=HTMLResponse)
async def serve_index():
    try:
        with open("index.html", "r", encoding="utf-8") as f:
            return f.read()
    except:
        return HTMLResponse("<h1>Welcome to ERA 75th Anniversary Registration</h1>")

@app.get("/{filename}.html", response_class=HTMLResponse)
async def serve_html(filename: str):
    file_path = f"{filename}.html"
    if os.path.exists(file_path):
        with open(file_path, "r", encoding="utf-8") as f:
            return f.read()
    raise HTTPException(status_code=404, detail="Page not found")

@app.get("/{filename}.png")
async def serve_png(filename: str):
    file_path = f"{filename}.png"
    if os.path.exists(file_path):
        return FileResponse(file_path, media_type="image/png")
    raise HTTPException(status_code=404, detail="Image not found")

# ============ API ROUTES ============
@app.get("/api/health")
def health(db: Session = Depends(get_db)):
    try:
        db.execute(text("SELECT 1"))
        return {"status": "healthy", "database": "connected", "type": "postgresql"}
    except Exception as e:
        return {"status": "healthy", "database": "error", "message": str(e)}

@app.post("/api/register")
def register(data: RegistrationCreate, db: Session = Depends(get_db)):
    try:
        # Check if email exists
        existing = db.query(Registration).filter(Registration.email == data.email).first()
        if existing:
            raise HTTPException(status_code=400, detail="Email already registered")
        
        # Generate guest ID
        guest_id = generate_guest_id()
        
        # Parse date of birth if provided
        dob = None
        if data.date_of_birth:
            try:
                dob = datetime.strptime(data.date_of_birth, '%Y-%m-%d').date()
            except:
                pass
        
        # Create new registration
        new_registration = Registration(
            guest_id=guest_id,
            full_name=data.full_name,
            gender=data.gender,
            phone=data.phone,
            email=data.email,
            date_of_birth=dob,
            organization=data.organization,
            job_title=data.job_title,
            city=data.city,
            category=data.category,
            registered_at=datetime.utcnow()
        )
        
        db.add(new_registration)
        db.commit()
        db.refresh(new_registration)
        
        # Send SMS notification
        sms_message = f"ERA75: Thank you {data.full_name}! Your Guest ID: {guest_id}. Show this at the gate. ERA 75th Anniversary."
        sms_result = send_sms(data.phone, sms_message)
        
        return {
            "guest_id": guest_id,
            "full_name": data.full_name,
            "email": data.email,
            "category": data.category,
            "message": "Registration successful",
            "sms": sms_result
        }
    except HTTPException as e:
        raise e
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/guests")
def get_guests(db: Session = Depends(get_db)):
    try:
        guests = db.query(Registration).order_by(Registration.registered_at.asc()).all()
        return [{
            "guest_id": g.guest_id,
            "full_name": g.full_name,
            "gender": g.gender,
            "phone": g.phone,
            "email": g.email,
            "date_of_birth": g.date_of_birth.isoformat() if g.date_of_birth else None,
            "organization": g.organization,
            "job_title": g.job_title,
            "city": g.city,
            "category": g.category,
            "registered_at": g.registered_at.isoformat() if g.registered_at else None,
            "checked_in": g.checked_in,
            "checked_in_at": g.checked_in_at.isoformat() if g.checked_in_at else None,
            "sms_sent": g.sms_sent
        } for g in guests]
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/stats")
def get_stats(db: Session = Depends(get_db)):
    try:
        total = db.query(Registration).count()
        checked_in = db.query(Registration).filter(Registration.checked_in == 1).count()
        return {
            "total": total,
            "checked_in": checked_in,
            "pending": total - checked_in,
            "check_in_rate": round((checked_in / total * 100) if total > 0 else 0, 1)
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# ============ BULK DELETE ENDPOINT ============
@app.post("/api/guests/bulk-delete")
def bulk_delete(guest_ids: List[str], db: Session = Depends(get_db)):
    try:
        if not guest_ids or len(guest_ids) == 0:
            raise HTTPException(status_code=400, detail="No guest IDs provided")
        
        deleted_count = 0
        for guest_id in guest_ids:
            guest = db.query(Registration).filter(Registration.guest_id == guest_id).first()
            if guest:
                db.delete(guest)
                deleted_count += 1
        
        db.commit()
        return {"message": f"Deleted {deleted_count} guests successfully"}
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))

# ============ EXPORT FUNCTIONS ============
@app.get("/api/export/csv")
def export_csv(db: Session = Depends(get_db)):
    try:
        guests = db.query(Registration).order_by(Registration.registered_at.asc()).all()
        
        output = StringIO()
        writer = csv.writer(output)
        
        writer.writerow([
            '#', 'Guest ID', 'Full Name', 'Gender', 'Phone', 'Email', 'Date of Birth',
            'Organization', 'Job Title', 'City', 'Category', 
            'Registered At', 'Checked In', 'Checked In At'
        ])
        
        for idx, guest in enumerate(guests, 1):
            writer.writerow([
                idx,
                guest.guest_id or '',
                guest.full_name or '',
                guest.gender or '',
                guest.phone or '',
                guest.email or '',
                guest.date_of_birth.isoformat() if guest.date_of_birth else '',
                guest.organization or '',
                guest.job_title or '',
                guest.city or '',
                guest.category or '',
                guest.registered_at.isoformat() if guest.registered_at else '',
                'Yes' if guest.checked_in == 1 else 'No',
                guest.checked_in_at.isoformat() if guest.checked_in_at else ''
            ])
        
        csv_content = output.getvalue()
        output.close()
        
        return Response(
            content=csv_content,
            media_type="text/csv",
            headers={
                "Content-Disposition": "attachment; filename=registrations_export.csv"
            }
        )
    except Exception as e:
        return Response(
            content=f"Error: {str(e)}",
            media_type="text/plain",
            status_code=500
        )

@app.get("/api/export/json")
def export_json(db: Session = Depends(get_db)):
    try:
        guests = db.query(Registration).order_by(Registration.registered_at.asc()).all()
        return [{
            "row_number": idx,
            "guest_id": g.guest_id,
            "full_name": g.full_name,
            "gender": g.gender,
            "phone": g.phone,
            "email": g.email,
            "date_of_birth": g.date_of_birth.isoformat() if g.date_of_birth else None,
            "organization": g.organization,
            "job_title": g.job_title,
            "city": g.city,
            "category": g.category,
            "registered_at": g.registered_at.isoformat() if g.registered_at else None,
            "checked_in": g.checked_in,
            "checked_in_at": g.checked_in_at.isoformat() if g.checked_in_at else None
        } for idx, g in enumerate(guests, 1)]
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/export/excel")
def export_excel(db: Session = Depends(get_db)):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        import io
        
        guests = db.query(Registration).order_by(Registration.registered_at.asc()).all()
        
        wb = Workbook()
        ws = wb.active
        ws.title = "ERA 75th Registrations"
        
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="FF6B00", end_color="FF6B00", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        headers = [
            '#', 'Guest ID', 'Full Name', 'Gender', 'Phone', 'Email', 'Date of Birth',
            'Organization', 'Job Title', 'City', 'Category', 
            'Registered At', 'Checked In', 'Checked In At'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        for row_idx, guest in enumerate(guests, 2):
            data = [
                row_idx - 1,
                guest.guest_id or '',
                guest.full_name or '',
                guest.gender or '',
                guest.phone or '',
                guest.email or '',
                guest.date_of_birth.isoformat() if guest.date_of_birth else '',
                guest.organization or '',
                guest.job_title or '',
                guest.city or '',
                guest.category or '',
                guest.registered_at.isoformat() if guest.registered_at else '',
                'Yes' if guest.checked_in == 1 else 'No',
                guest.checked_in_at.isoformat() if guest.checked_in_at else ''
            ]
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.border = border
                cell.alignment = Alignment(horizontal="left", vertical="center")
                if col == 13:
                    if value == 'Yes':
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    else:
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        for col in range(1, len(headers) + 1):
            column_letter = get_column_letter(col)
            ws.column_dimensions[column_letter].auto_size = True
            if col in [3, 6, 7, 8]:
                ws.column_dimensions[column_letter].width = 25
        
        ws.freeze_panes = 'A2'
        
        summary_ws = wb.create_sheet("Summary")
        summary_ws['A1'] = "ERA 75th Anniversary Event"
        summary_ws['A1'].font = Font(bold=True, size=16, color="FF6B00")
        summary_ws['A3'] = "Registration Summary Report"
        summary_ws['A3'].font = Font(bold=True, size=14)
        
        total = len(guests)
        checked_in = sum(1 for g in guests if g.checked_in == 1)
        pending = total - checked_in
        
        summary_data = [
            ['', ''],
            ['Total Registrations', total],
            ['Checked In', checked_in],
            ['Pending', pending],
            ['Check-in Rate', f"{round((checked_in/total*100) if total > 0 else 0, 1)}%"],
            ['', ''],
            ['Generated On', datetime.now().strftime('%Y-%m-%d %H:%M:%S')]
        ]
        
        for row_idx, (label, value) in enumerate(summary_data, 5):
            summary_ws.cell(row=row_idx, column=1, value=label)
            summary_ws.cell(row=row_idx, column=2, value=value)
            if row_idx > 5:
                summary_ws.cell(row=row_idx, column=1).font = Font(bold=True)
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return Response(
            content=output.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": "attachment; filename=registrations_export.xlsx"
            }
        )
    except Exception as e:
        return Response(
            content=f"Error: {str(e)}",
            media_type="text/plain",
            status_code=500
        )

# ============ CERTIFICATE GENERATION ============
@app.get("/api/certificate/{guest_id}")
def generate_certificate(guest_id: str, db: Session = Depends(get_db)):
    try:
        guest = db.query(Registration).filter(Registration.guest_id == guest_id).first()
        if not guest:
            raise HTTPException(status_code=404, detail="Guest not found")
        
        buffer = BytesIO()
        page_width, page_height = landscape(A4)
        c = canvas.Canvas(buffer, pagesize=landscape(A4))
        
        # Background
        c.setFillColorRGB(0.98, 0.96, 0.92)
        c.rect(0, 0, page_width, page_height, fill=1, stroke=0)
        
        # Border
        c.setStrokeColor(HexColor("#FF6B00"))
        c.setLineWidth(4)
        c.rect(20, 20, page_width - 40, page_height - 40)
        
        c.setStrokeColor(HexColor("#D4AF37"))
        c.setLineWidth(1.5)
        c.rect(30, 30, page_width - 60, page_height - 60)
        
        # Corner decorations
        corner_size = 30
        c.setStrokeColor(HexColor("#FF6B00"))
        c.setLineWidth(2)
        c.line(35, page_height - 35, 35 + corner_size, page_height - 35)
        c.line(35, page_height - 35, 35, page_height - 35 - corner_size)
        c.line(page_width - 35, page_height - 35, page_width - 35 - corner_size, page_height - 35)
        c.line(page_width - 35, page_height - 35, page_width - 35, page_height - 35 - corner_size)
        c.line(35, 35, 35 + corner_size, 35)
        c.line(35, 35, 35, 35 + corner_size)
        c.line(page_width - 35, 35, page_width - 35 - corner_size, 35)
        c.line(page_width - 35, 35, page_width - 35, 35 + corner_size)
        
        # Watermark
        try:
            watermark_path = "era_logo.png"
            if os.path.exists(watermark_path):
                watermark = ImageReader(watermark_path)
                c.saveState()
                c.setFillColorRGB(1, 1, 1, 0.08)
                c.drawImage(watermark, page_width/2 - 150, page_height/2 - 150, 
                           width=300, height=300, preserveAspectRatio=True, mask='auto')
                c.restoreState()
        except:
            pass
        
        # Logos
        logo_y = page_height - 75
        
        try:
            left_logo_path = "era_75_logo.png"
            if os.path.exists(left_logo_path):
                left_logo = ImageReader(left_logo_path)
                c.drawImage(left_logo, 50, logo_y - 35, width=80, height=60, preserveAspectRatio=True)
        except:
            pass
        
        try:
            right_logo_path = "era_logo.png"
            if os.path.exists(right_logo_path):
                right_logo = ImageReader(right_logo_path)
                c.drawImage(right_logo, page_width - 130, logo_y - 35, width=80, height=60, preserveAspectRatio=True)
        except:
            pass
        
        # Title
        title_y = page_height - 85
        c.setFillColor(HexColor("#D4AF37"))
        c.setFont("Helvetica-Bold", 30)
        c.drawCentredString(page_width / 2, title_y, "CERTIFICATE OF PARTICIPATION")
        
        c.setStrokeColor(HexColor("#D4AF37"))
        c.setLineWidth(2)
        c.line(page_width / 2 - 180, title_y - 15, page_width / 2 + 180, title_y - 15)
        
        # Subtitle
        c.setFillColor(HexColor("#555555"))
        c.setFont("Helvetica", 13)
        c.drawCentredString(page_width / 2, title_y - 42, "This certificate is proudly presented to")
        
        # Guest Name
        c.setFillColor(HexColor("#FF6B00"))
        c.setFont("Helvetica-Bold", 38)
        name = guest.full_name.upper() if guest.full_name else ""
        c.drawCentredString(page_width / 2, title_y - 100, name)
        
        c.setStrokeColor(HexColor("#FF6B00"))
        c.setLineWidth(2)
        c.line(page_width / 2 - 160, title_y - 115, page_width / 2 + 160, title_y - 115)
        
        # Main Text
        c.setFillColor(HexColor("#333333"))
        c.setFont("Helvetica", 13)
        c.drawCentredString(page_width / 2, title_y - 145, "for their valuable participation in the")
        c.drawCentredString(page_width / 2, title_y - 168, "ERA 75th Anniversary Event")
        
        # Guest Details
        detail_y = title_y - 220
        
        c.setFillColor(HexColor("#555555"))
        c.setFont("Helvetica", 11)
        c.drawString(page_width / 2 - 180, detail_y, "Guest ID:")
        c.setFillColor(HexColor("#FF6B00"))
        c.setFont("Helvetica-Bold", 13)
        c.drawString(page_width / 2 - 100, detail_y, guest.guest_id or '')
        
        c.setFillColor(HexColor("#555555"))
        c.setFont("Helvetica", 11)
        c.drawString(page_width / 2 - 180, detail_y - 22, "Organization:")
        c.setFillColor(HexColor("#333333"))
        c.setFont("Helvetica", 12)
        c.drawString(page_width / 2 - 100, detail_y - 22, guest.organization or 'N/A')
        
        c.setFillColor(HexColor("#555555"))
        c.setFont("Helvetica", 11)
        c.drawString(page_width / 2 - 180, detail_y - 44, "Category:")
        c.setFillColor(HexColor("#333333"))
        c.setFont("Helvetica", 12)
        c.drawString(page_width / 2 - 100, detail_y - 44, guest.category or '')
        
        # Date
        c.setFillColor(HexColor("#555555"))
        c.setFont("Helvetica", 11)
        c.drawCentredString(page_width / 2, detail_y - 75, "Date: July 23, 2026")
        
        # Signature
        signature_y = 120
        
        c.setStrokeColor(HexColor("#333333"))
        c.setLineWidth(1)
        c.line(page_width / 2 - 100, signature_y, page_width / 2 + 100, signature_y)
        
        try:
            signature_path = "director_signature.png"
            if os.path.exists(signature_path):
                signature = ImageReader(signature_path)
                c.drawImage(signature, page_width / 2 - 60, signature_y + 5, 
                           width=120, height=40, preserveAspectRatio=True, mask='auto')
            else:
                c.setFillColor(HexColor("#333333"))
                c.setFont("Helvetica", 11)
                c.drawCentredString(page_width / 2, signature_y + 10, "Director General")
        except:
            c.setFillColor(HexColor("#333333"))
            c.setFont("Helvetica", 11)
            c.drawCentredString(page_width / 2, signature_y + 10, "Director General")
        
        c.setFillColor(HexColor("#555555"))
        c.setFont("Helvetica", 10)
        c.drawCentredString(page_width / 2, signature_y - 18, "Director General")
        c.drawCentredString(page_width / 2, signature_y - 32, "Ethiopian Roads Administration")
        
        # Footer divider
        c.setStrokeColor(HexColor("#D4AF37"))
        c.setLineWidth(0.5)
        c.line(page_width / 2 - 60, signature_y - 50, page_width / 2 + 60, signature_y - 50)
        
        # Footer
        c.setFillColor(HexColor("#888888"))
        c.setFont("Helvetica", 8)
        c.drawCentredString(page_width / 2, 30, "Ethiopian Roads Administration - ERA 75th Anniversary Celebration")
        c.drawCentredString(page_width / 2, 18, f"Generated on {datetime.now().strftime('%B %d, %Y at %H:%M')}")
        
        c.save()
        buffer.seek(0)
        
        return Response(
            content=buffer.getvalue(),
            media_type="application/pdf",
            headers={
                "Content-Disposition": f"attachment; filename=certificate_{guest_id}.pdf"
            }
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Certificate generation failed: {str(e)}")

# ============ CHECK-IN ============
@app.post("/api/checkin/{guest_id}")
def check_in(guest_id: str, db: Session = Depends(get_db)):
    try:
        guest = db.query(Registration).filter(Registration.guest_id == guest_id).first()
        if not guest:
            raise HTTPException(status_code=404, detail="Guest not found")
        
        if guest.checked_in == 1:
            raise HTTPException(status_code=400, detail="Guest already checked in")
        
        guest.checked_in = 1
        guest.checked_in_at = datetime.utcnow()
        db.commit()
        
        return {"message": "Guest checked in successfully", "guest_id": guest_id}
    except HTTPException as e:
        raise e
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))

# ============ ANALYTICS ============
@app.get("/api/analytics/daily")
def get_daily_stats(db: Session = Depends(get_db)):
    try:
        from sqlalchemy import func, cast, Date
        thirty_days_ago = datetime.utcnow() - timedelta(days=30)
        results = db.query(
            cast(Registration.registered_at, Date).label('date'),
            func.count().label('count')
        ).filter(Registration.registered_at >= thirty_days_ago)\
         .group_by(cast(Registration.registered_at, Date))\
         .order_by('date').all()
        
        return [{"date": str(r.date), "count": r.count} for r in results]
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/analytics/categories")
def get_category_breakdown(db: Session = Depends(get_db)):
    try:
        results = db.query(
            Registration.category,
            func.count().label('count')
        ).group_by(Registration.category).order_by(func.count().desc()).all()
        
        return [{"category": r.category, "count": r.count} for r in results]
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/analytics/checkins")
def get_checkin_analytics(db: Session = Depends(get_db)):
    try:
        from sqlalchemy import func, extract
        total = db.query(Registration).count()
        checked = db.query(Registration).filter(Registration.checked_in == 1).count()
        
        results = db.query(
            extract('hour', Registration.checked_in_at).label('hour'),
            func.count().label('count')
        ).filter(Registration.checked_in == 1)\
         .filter(Registration.checked_in_at.isnot(None))\
         .group_by(extract('hour', Registration.checked_in_at))\
         .order_by('hour').all()
        
        return {
            "total_guests": total,
            "checked_in": checked,
            "pending": total - checked,
            "check_in_rate": round((checked / total * 100) if total > 0 else 0, 1),
            "by_hour": [{"hour": str(int(r.hour)).zfill(2), "count": r.count} for r in results]
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# ============ SEARCH ============
@app.get("/api/search")
def search_guests(q: str, db: Session = Depends(get_db)):
    try:
        query = f"%{q}%"
        guests = db.query(Registration).filter(
            Registration.full_name.like(query) |
            Registration.email.like(query) |
            Registration.guest_id.like(query) |
            Registration.phone.like(query) |
            Registration.organization.like(query)
        ).order_by(Registration.registered_at.asc()).all()
        
        return [{
            "guest_id": g.guest_id,
            "full_name": g.full_name,
            "email": g.email,
            "phone": g.phone,
            "organization": g.organization,
            "category": g.category,
            "registered_at": g.registered_at.isoformat() if g.registered_at else None,
            "checked_in": g.checked_in
        } for g in guests]
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# ============ SMS TEST ENDPOINT ============
@app.get("/api/test-sms")
def test_sms():
    """Test endpoint to verify SMS configuration"""
    test_phone = os.environ.get('TEST_PHONE_NUMBER', '0912345678')
    test_message = "ERA75 Test: Your SMS is working!"
    result = send_sms(test_phone, test_message)
    return {
        "message": "SMS test completed",
        "phone": test_phone,
        "result": result,
        "note": "If simulated, no actual SMS was sent. To enable real SMS, set TWILIO credentials."
    }

# ============ RUN ============
if __name__ == "__main__":
    import uvicorn
    port = int(os.environ.get("PORT", 8000))
    uvicorn.run(app, host="0.0.0.0", port=port)
